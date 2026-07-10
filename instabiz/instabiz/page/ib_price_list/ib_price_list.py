import re
import os
import frappe
from frappe.utils import flt, today, now_datetime


_EXCEL_PATH = os.path.join(
	os.path.dirname(__file__),
	"../../../../scripts/PRICE LIST.xlsx",
)

_PRICE_FIELDS = {"face_price", "last_price", "slab1", "slab2", "slab3", "slab4", "slab5"}


@frappe.whitelist()
def get_rate_card_meta():
	"""Return summary counts + last effective dates per product type."""
	rows = frappe.db.sql("""
		SELECT product_type,
		       COUNT(*) AS cnt,
		       MAX(effective_date) AS last_date
		FROM `tabIB Rate Card Entry`
		GROUP BY product_type
	""", as_dict=1)
	return {r.product_type: {"count": r.cnt, "last_date": str(r.last_date or "")} for r in rows}


@frappe.whitelist()
def reimport_rate_card():
	"""Re-import all entries from the canonical PRICE LIST.xlsx in scripts/."""
	frappe.only_for(["System Manager", "Sales Manager"])
	if not os.path.exists(_EXCEL_PATH):
		frappe.throw(f"File not found: {_EXCEL_PATH}")

	import openpyxl
	wb = openpyxl.load_workbook(_EXCEL_PATH, data_only=True)

	def safe(v):
		return (str(v).strip() if v is not None else "")

	jumbo_rows, cut_rows = [], []

	ws1 = wb["PRICE LIST JUMBO 1.5.26"]
	for row in ws1.iter_rows(min_row=10, max_row=ws1.max_row, values_only=True):
		code = safe(row[1])
		if not code.startswith("IS-"):
			continue
		face, last = row[8], row[9]
		if face is None and last is None:
			continue
		jumbo_rows.append({
			"product_type": "Jumbo Roll",
			"item_code": code,
			"product_name": safe(row[2]),
			"color": safe(row[3]),
			"liner": safe(row[4]),
			"thickness": safe(row[5]),
			"width_mm": row[6],
			"length_m": row[7],
			"face_price": face or 0,
			"last_price": last or 0,
			"slab1": face or 0,
			"slab5": last or 0,
			"unit": safe(row[10]),
			"effective_date": "2026-05-20",
		})

	ws2 = wb["PRICE LIST CUT PACK 1.5.26"]
	for row in ws2.iter_rows(min_row=10, max_row=ws2.max_row, values_only=True):
		code = safe(row[1])
		if not code.startswith("IS-"):
			continue
		s1, s2, s3, s4, s5 = row[8], row[9], row[10], row[11], row[12]
		if not any(v is not None for v in [s1, s2, s3, s4, s5]):
			continue
		cut_rows.append({
			"product_type": "Cut Pack",
			"item_code": code,
			"product_name": safe(row[2]),
			"color": safe(row[3]),
			"liner": safe(row[4]),
			"thickness": safe(row[5]),
			"width_mm": row[6],
			"length_m": row[7],
			"slab1": s1 or 0,
			"slab2": s2 or 0,
			"slab3": s3 or 0,
			"slab4": s4 or 0,
			"slab5": s5 or 0,
			"face_price": s1 or 0,
			"last_price": s5 or 0,
			"unit": safe(row[13]),
			"packing_standard": safe(row[14]),
			"effective_date": today(),
		})

	frappe.db.delete("IB Rate Card Entry", {})
	for r in jumbo_rows + cut_rows:
		frappe.get_doc({"doctype": "IB Rate Card Entry", **r}).insert(ignore_permissions=True)
	frappe.db.commit()

	return {
		"jumbo": len(jumbo_rows),
		"cut_pack": len(cut_rows),
		"total": len(jumbo_rows) + len(cut_rows),
	}


@frappe.whitelist()
def get_rate_card(product_type=None):
	"""Return all IB Rate Card Entry rows, optionally filtered by product_type."""
	filters = {}
	if product_type:
		filters["product_type"] = product_type
	rows = frappe.db.get_all(
		"IB Rate Card Entry",
		filters=filters,
		fields=[
			"name", "product_type", "item_code", "product_name",
			"color", "liner", "thickness", "width_mm", "length_m", "unit",
			"face_price", "last_price",
			"slab1", "slab2", "slab3", "slab4", "slab5",
			"packing_standard", "effective_date",
		],
		order_by="item_code asc, product_name asc",
	)
	for r in rows:
		w = _fmt_dim(r.get("width_mm"))
		l = _fmt_dim(r.get("length_m"))
		r["spec_dimension"] = f"{w}mm × {l}m" if (w and l) else ""
		r["spec_sqm"] = round(flt(w) / 1000 * flt(l), 3) if (w and l) else ""
	return rows


@frappe.whitelist()
def save_rate_card_entry(name, data):
	"""Update an existing IB Rate Card Entry. Managers only."""
	frappe.only_for(["System Manager", "Sales Manager"])
	if isinstance(data, str):
		data = frappe.parse_json(data)
	doc = frappe.get_doc("IB Rate Card Entry", name)
	allowed = {
		"item_code", "product_name", "color", "liner", "thickness",
		"width_mm", "length_m", "unit", "effective_date", "packing_standard",
		"face_price", "last_price", "slab1", "slab2", "slab3", "slab4", "slab5",
	}
	for k, v in data.items():
		if k in allowed:
			setattr(doc, k, v if v != "" else None)
	# keep face/last in sync with slabs for Cut Pack
	if doc.product_type == "Cut Pack":
		doc.face_price = doc.slab1 or 0
		doc.last_price = doc.slab5 or 0
	doc.save(ignore_permissions=True)
	_notify_price_change(doc, "update")
	return {"name": doc.name}


@frappe.whitelist()
def add_rate_card_entry(data):
	"""Create a new IB Rate Card Entry. Managers only."""
	frappe.only_for(["System Manager", "Sales Manager"])
	if isinstance(data, str):
		data = frappe.parse_json(data)
	if data.get("product_type") == "Cut Pack":
		data["face_price"] = data.get("slab1") or 0
		data["last_price"] = data.get("slab5") or 0
	doc = frappe.get_doc({"doctype": "IB Rate Card Entry", **data})
	doc.insert(ignore_permissions=True)
	frappe.db.commit()
	_notify_price_change(doc, "add")
	w = _fmt_dim(doc.get("width_mm"))
	l = _fmt_dim(doc.get("length_m"))
	result = doc.as_dict()
	result["spec_dimension"] = f"{w}mm × {l}m" if (w and l) else ""
	return result


@frappe.whitelist()
def delete_rate_card_entry(name):
	"""Delete an IB Rate Card Entry. System Manager only."""
	frappe.only_for(["System Manager"])
	frappe.delete_doc("IB Rate Card Entry", name, ignore_permissions=True)
	frappe.db.commit()
	return {"ok": 1}


@frappe.whitelist()
def get_price_history(name):
	"""Return chronological price-change log for an IB Rate Card Entry."""
	versions = frappe.db.get_all(
		"Version",
		filters={"ref_doctype": "IB Rate Card Entry", "docname": name},
		fields=["creation", "owner", "data"],
		order_by="creation desc",
		limit=100,
	)
	history = []
	for v in versions:
		try:
			vdata = frappe.parse_json(v.data or "{}")
		except Exception:
			continue
		changed = [c for c in (vdata.get("changed") or []) if c[0] in _PRICE_FIELDS]
		if not changed:
			continue
		history.append({
			"timestamp": str(v.creation),
			"user": v.owner,
			"changes": [[c[0], _parse_num(c[1]), _parse_num(c[2])] for c in changed],
		})
	return history


def _notify_price_change(doc, action):
	"""Publish realtime event + Notification Log for all sales-role users."""
	try:
		me = frappe.session.user
		item_label = doc.item_code or doc.name
		product_label = doc.product_name or item_label

		subject = (
			f"New price list item: {item_label}" if action == "add"
			else f"Price updated: {item_label}"
		)
		body = (
			f"{product_label} ({doc.product_type}) "
			+ ("was added to the price list" if action == "add" else "prices were updated")
		)

		# Users in sales roles (excluding the one who made the change)
		sales_users = frappe.db.sql("""
			SELECT DISTINCT ur.parent AS user
			FROM   `tabHas Role` ur
			INNER JOIN `tabUser` u ON u.name = ur.parent AND u.enabled = 1
			WHERE  ur.role IN ('Sales Manager','Sales User','Sales Executive')
			  AND  ur.parenttype = 'User'
			  AND  ur.parent != %(me)s
		""", {"me": me}, as_dict=1)

		for row in sales_users:
			frappe.get_doc({
				"doctype":       "Notification Log",
				"subject":       subject,
				"email_content": body,
				"for_user":      row.user,
				"document_type": "IB Rate Card Entry",
				"document_name": doc.name,
				"type":          "Alert",
				"from_user":     me,
			}).insert(ignore_permissions=True)

		# Realtime push — reaches every connected browser (not per-user)
		frappe.publish_realtime(
			"ib_price_list_updated",
			{
				"action":       action,        # "add" | "update"
				"item_code":    doc.item_code,
				"product_name": doc.product_name or "",
				"product_type": doc.product_type or "",
				"changed_by":   frappe.utils.get_fullname(me),
			},
			after_commit=True,
		)
	except Exception:
		frappe.log_error(frappe.get_traceback(), "IB price list notify failed")


def _parse_num(val):
	if val is None:
		return 0.0
	cleaned = re.sub(r"[^\d.\-]", "", str(val))
	return flt(cleaned) if cleaned else 0.0


def _fmt_dim(val):
	if not val:
		return None
	f = flt(val)
	return str(int(f)) if f == int(f) else str(f)
